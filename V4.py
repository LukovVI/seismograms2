import os
import cv2
import pandas as pd
import openpyxl



def rol(y, x, img):#подсчёт чёрных соседей по прямым направлениям
    p = 0
    for i, j in list(([1, 0], [0, -1], [-1, 0], [0, 1])):
        if img[y + i, x + j] in [0, 150]:
            p += 1
    return p

def smallchek(y, x, img, threshold):#поиск точки для старта одхода
    dx = 10
    dy = 15
    if dx*dy - cv2.countNonZero(img[y:y+dy, x:x+dx]) < threshold:#проверка маленького крвдрата на плотность
        return False
    for i in range(y, y+dy):
        for j in range(x, x+dx):
            if img[i, j] == 0:
                if rol(i, j, img) >= 3:#у точки должно быть от 5-ти соседей, но проверки прямых направлений на наличие 3-х соседей достаточно
                    return list([i, j])
    return False

def drawer(l, img):#закраска зоны калибровки
    for q in l:
        img[q[0], q[1]] = 200

def plot2(img, t):#измерение плотности 2-х часового промежутка
    black2 = (img[t*330:(t+1)*330, 81:3989] == 0).sum() + (img[t*330:(t+1)*330, 81:3989] == 100).sum()
    all2 = (img[t*330:(t+1)*330, 81:3989] == 255).sum() + black2
    return black2/all2, all2, black2


def plot12(img):#массив из плотностей по 2 часа по порядку, последний элемент средняя плотность всего изображения
    plot_12_2 = []
    black = 0
    all = 0
    for i in range(6):
        pl2 = plot2(img, i)
        plot_12_2.append(pl2[0])
        all += pl2[1]
        black += pl2[2]
    plot_12_2.append(black/all)
    return plot_12_2

def bfs(y, x, img):#обход в ширину для поиска всех точек принадлежащих кластеру
    l = [[y, x]]#список подозрительных на принадлежность к карректировке точек
    visit = {y: [x]}#словарь посещённых точек для предотвращения цикла (словарь для большей скорости)
    check = [[y, x]]#точки подлежащие посещению
    while check:
        y, x = check.pop()
        if rol(y, x, img) >= 3:
            for i, j in list(([1, 0], [0, -1], [-1, 0], [0, 1])):
                if y+i not in visit:
                    k = {y+i: []}
                    visit.update(k)
                if img[y + i, x + j] == 0 and x+j not in visit[y+i]:
                    visit[y+i].append(x+j)
                    check.append([y+i, x+j])
            img[y, x] = 100
            l.append([y, x])
    return [len(l), l]

def all(name):
    my_place = os.getcwd() + "\\"
    image = cv2.imread(my_place + name + ".png", 0)
    height, width = image.shape[:2]
    # plot = [name]# подсчёт плотности до обработки
    # plot.extend(plot12(image))
    # time = list(["name", "0-2", "2-4", "4-6", "6-8", "8-10", "10-12", "0-12"])
    # print(time[0] + "\t\t" + plot[0])
    # for p in range(1, 8):
    #     print(time[p] + "\t\t{0:.5f}".format(plot[p]))
    # print()
    threshold = 135
    for i in range(5, height - 10, 5):#проход по отдельным квадратам
        for j in range(5, width - 10, 5):
            check_point = smallchek(i, j, image, threshold)
            if check_point:
                le, l = bfs(check_point[0], check_point[1], image)
                if le >= 200:
                    # print(le, end = " ") #вывод размера кластера
                    drawer(l, image)
    for i in range(80, 3991, 782):#удаление вертикальных полос
        for j in range(0, 2000):
            image[j, i] = 200

    #создание папки с обработанными файлами
    fin = my_place + "final\\"
    if not os.path.exists(fin):
        os.mkdir(fin)
    cv2.imwrite(fin + name + "_final.png", image)

    #вывод результатов в консоль
    print()
    plot = [name]
    plot.extend(plot12(image))
    time = list(["name", "0-2", "2-4", "4-6", "6-8", "8-10", "10-12", "0-12"])
    print(time[0] + "\t\t" + plot[0])
    for p in range(1, 8):
        print(time[p] + "\t\t{0:.5f}".format(plot[p]))
    print()

    #вывод результатов в excel файл
    if not os.path.exists(my_place + "BD.xlsx"):
        columns = {}
        colomns_up = [[time[i], []] for i in range(8)]
        columns.update(colomns_up)
        pd.DataFrame(columns).to_excel(my_place + "BD.xlsx", index=False)
    wbb = openpyxl.load_workbook(filename = 'BD.xlsx')
    wb = wbb.active
    wr = wb.max_row + 1
    app = [name]
    app.extend(["%.5f" % i for i in plot[1:]])
    for col in range(8):
        wb.cell(row=wr, column=col+1).value = app[col]
    wbb.save(my_place + "BD.xlsx")


def main():#проход по всем файлам
    for i in list(map(lambda x: x[0:-4], list(filter(lambda x: x.endswith(".png"), os.listdir())))):
        all(i)
    input("нажмине Enter для закрытия консоли")

if __name__ == '__main__':
    main()