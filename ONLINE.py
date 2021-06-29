import os
import sys
import cv2#pip install opencv-python
import pandas as pd#pip install pandas
import openpyxl#pip install openpyxl
import numpy as np#pip install numpy
import requests#pip install requests
import datetime
#для преобразования в .exe файл прописать в командной строке pyinstaller proj_disp.py -F



def rol(y, x, img):#подсчёт чёрных соседей по прямым направлениям
    p = 0
    for i, j in list(([1, 0], [0, -1], [-1, 0], [0, 1])):
        if img[y + i, x + j] in [0, 100]:
            p += 1
    return p


def smallchek(y, x, img, threshold):#поиск точки для старта одхода
    dx = 10
    dy = 15
    if dx*dy - cv2.countNonZero(img[y:y+dy, x:x+dx]) < threshold:#проверка маленького крвдрата на плотность
        return False
    for i in range(y, y+dy):
        for j in range(x, x+dx):
            if img[i, j] == 0:
                if rol(i, j, img) >= 3:#у точки должно быть от 5-ти соседей, но проверки прямых направлений на наличие 3-х соседей достаточно
                    return list([i, j])
    return False


def drawer(l, img):#закраска зоны калибровки
    for q in l:
        img[q[0], q[1]] = 200


def plot2(img, t):#измерение плотности 2-х часового промежутка
    black2 = (img[t:t+330, 81:3989] == 0).sum() + (img[t:t+330, 81:3989] == 100).sum()
    all2 = (img[t:t+330, 81:3989] == 255).sum() + black2
    return black2/all2, all2, black2


def plot12(img):#массив из плотностей по 2 часа по порядку, последний элемент средняя плотность всего изображения
    plot_12_2 = []
    black = 0
    all = 0
    for i in range(6):
        pl2 = plot2(img, i*330)
        plot_12_2.append(pl2[0])
        all += pl2[1]
        black += pl2[2]
    plot_12_2.append(black/all)
    return plot_12_2


def lower_line(img):
    b = 4000
    for x in range(81, 3980, 5):
        for y in range(1989, 1800, -5):
            if lower_line_small_check(x, y, img):
                if y<b:
                    b = y
            else:
                break
    return b


def lower_line_small_check(x, y, img):
    dx = 5
    dy = 5
    if cv2.countNonZero(img[y:y + dy, x:x + dx]) == dx*dy:
        return True
    return False


def bfs(y, x, img):#обход в ширину для поиска всех точек принадлежащих кластеру
    l = [[y, x]]#список подозрительных на принадлежность к карректировке точек
    visit = {y: [x]}#словарь посещённых точек для предотвращения цикла (словарь для большей скорости)
    check = [[y, x]]#точки подлежащие посещению
    while check:
        y, x = check.pop()
        if y > 0 and y < 1999:
            if rol(y, x, img) >= 3:
                for i, j in list(([1, 0], [0, -1], [-1, 0], [0, 1])):
                    if y+i not in visit:
                        k = {y+i: []}
                        visit.update(k)
                    if img[y + i, x + j] == 0 and x+j not in visit[y+i]:
                        visit[y+i].append(x+j)
                        check.append([y+i, x+j])
                img[y, x] = 100
                l.append([y, x])
    return [len(l), l]


def full_check(img, height, width, start = 5):#проверка квадратов по всей облости начиная со start по высоте
    threshold = 135
    for i in range(start, height - 20, 5):
        for j in range(81, width - 10, 5):
            check_point = smallchek(i, j, img, threshold)
            if check_point:
                le, l = bfs(check_point[0], check_point[1], img)
                if le >= 200:
                    # print(le, end = " ") #вывод размера кластера
                    drawer(l, img)#закраска калибровки


def vert_bord(img):#удаление вертикальных полос
    for i in range(80, 3991, 782):
        for j in range(0, 2000):
            img[j, i] = 200


def online(url, place_excel, save_bd):
    resp = requests.get(url, stream=True).raw
    image = np.asarray(bytearray(resp.read()), dtype="uint8")
    image = cv2.imdecode(image, 0)
    height, width = image.shape[:2]
    y = lower_line(image)
    # for i in range(80, 4000, 2): #нижняя граница расчётов
    #     image[y:y+3, i] = 100

    full_check(image, height, width, y-450)#проверка картинки на наличие калибровки и её удаление

    vert_bord(image)#удаление вертикальных линий

    plot = "%.5f" % plot2(image, y-330)[0]
    print(plot)

    #вывод результатов в excel файл
    if save_bd:
        if not os.path.exists(place_excel + "small_BD_DONT_TOUCH.xlsx"):
            columns = {"time": [], "density": []}
            pd.DataFrame(columns).to_excel(place_excel + "small_BD_DONT_TOUCH.xlsx", index=False)
        wbb = openpyxl.load_workbook(filename = "small_BD_DONT_TOUCH.xlsx")
        wb = wbb.active
        wr = wb.max_row + 1
        wb.cell(row=wr, column=1).value = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        wb.cell(row=wr, column=2).value = plot
        try:
            wbb.save(place_excel + "small_BD_copy.xlsx")
        except:
            pass
        wbb.save(place_excel + "small_BD_DONT_TOUCH.xlsx")

    return plot


def all_monit(url, place_monit):
    resp = requests.get(url, stream=True).raw
    image = np.asarray(bytearray(resp.read()), dtype="uint8")
    image = cv2.imdecode(image, 0)
    height, width = image.shape[:2]
    name = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S_") + url.split("/")[-1][:-4]

    fin = place_monit + "base_monit\\"
    if not os.path.exists(fin):
        os.mkdir(fin)
    cv2.imwrite(fin + name + ".png", image)

    full_check(image, height, width)  # проверка картинки на наличие калибровки и её удаление

    vert_bord(image)  # удаление вертикальных линий

    # создание папки с обработанными файлами
    fin = place_monit + "final_monit\\"
    if not os.path.exists(fin):
        os.mkdir(fin)
    cv2.imwrite(fin + name + "_final.png", image)

    # вывод результатов в консоль
    print()
    plot = [name]
    plot.extend(plot12(image))
    time = list(["name", "0-2", "2-4", "4-6", "6-8", "8-10", "10-12", "0-12"])

    #вывод в консоль
    print(time[0] + "\t\t" + plot[0])
    for p in range(1, 8):
        print(time[p] + "\t\t{0:.5f}".format(plot[p]))
    print()

    # #вывод результатов в excel файл
    if not os.path.exists(place_monit + "BD_monit_DONT_TOUCH.xlsx"):
        columns = {}
        colomns_up = [[time[i], []] for i in range(8)]
        columns.update(colomns_up)
        pd.DataFrame(columns).to_excel(place_monit + "BD_monit_DONT_TOUCH.xlsx", index=False)
    wbb = openpyxl.load_workbook(filename='BD_monit_DONT_TOUCH.xlsx')
    wb = wbb.active
    wr = wb.max_row + 1
    app = [name]
    app.extend(["%.5f" % i for i in plot[1:]])
    for col in range(8):
        wb.cell(row=wr, column=col + 1).value = app[col]
    try:
        wbb.save(place_monit + "BD_monit_copy.xlsx")
    except:
        pass
    wbb.save(place_monit + "BD_monit_DONT_TOUCH.xlsx")

def all(name, dir, dir_save):
    image = cv2.imread(dir + name + ".png", 0)
    height, width = image.shape[:2]

    full_check(image, height, width)#проверка картинки на наличие калибровки и её удаление

    vert_bord(image)#удаление вертикальных линий

    #создание папки с обработанными файлами
    fin = dir_save + "final\\"
    if not os.path.exists(fin):
        os.mkdir(fin)
    cv2.imwrite(fin + name + "_final.png", image)

    #вывод результатов в консоль
    print()
    plot = [name]
    plot.extend(plot12(image))
    time = list(["name", "0-2", "2-4", "4-6", "6-8", "8-10", "10-12", "0-12"])
    print(time[0] + "\t\t" + plot[0])
    for p in range(1, 8):
        print(time[p] + "\t\t{0:.5f}".format(plot[p]))
    print()

    #вывод результатов в excel файл
    if not os.path.exists(dir_save + "BD_DONT_TOUCH.xlsx"):
        columns = {}
        colomns_up = [[time[i], []] for i in range(8)]
        columns.update(colomns_up)
        pd.DataFrame(columns).to_excel(dir_save + "BD_DONT_TOUCH.xlsx", index=False)
    wbb = openpyxl.load_workbook(filename = 'BD_DONT_TOUCH.xlsx')
    wb = wbb.active
    wr = wb.max_row + 1
    app = [name]
    app.extend(["%.5f" % i for i in plot[1:]])
    for col in range(8):
        wb.cell(row=wr, column=col+1).value = app[col]
    try:
        wbb.save(dir_save + "BD_copy.xlsx")
    except:
        pass
    wbb.save(dir_save + "BD_DONT_TOUCH.xlsx")

def main(linck, dir_save):#проход по всем файлам
    for name in list(map(lambda x: x[0:-4], list(filter(lambda x: x.endswith(".png"), os.listdir(linck))))):
        all(name, linck + "\\", dir_save)
